import csv
from openpyxl import load_workbook

def __search_by_column(column_name, file_path):
    if file_path.endswith('.csv'):
        return __search_by_first_value_in_column_csv(column_name, file_path)
    elif file_path.endswith('.xlsx') or file_path.endswith('.xls'):
        return __search_by_first_value_in_column_excel(column_name, file_path)
    else:
        raise ValueError("Unsupported file format. Please use a CSV or Excel file.")

def __search_by_first_value_in_column_csv(column_name, file_path):
    with open(file_path, mode='r', encoding='utf-8') as file:
        reader = csv.reader(file)
        headers = next(reader)  # 读取第一行（表头）

        if column_name in headers:
            column_index = headers.index(column_name)
            result = []

            for row in reader:
                result.append(row[column_index])
                
            return result
        else:
            print(f"未找到列首值为 '{column_name}' 的列。")
            return []

def __search_by_first_value_in_column_excel(column_name, file_path):
    workbook = load_workbook(filename=file_path, data_only=True)
    sheet = workbook.active  # 获取第一个工作表

    headers = [cell.value for cell in sheet[1]]  # 获取第一行的值作为表头

    if column_name in headers:
        column_index = headers.index(column_name) + 1
        result = []

        for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True):
            result.append(row[0])

        return result
    else:
        print(f"未找到列首值为 '{column_name}' 的列。")
        return []

# 示例用法：
column_name = 'Age'  # 你想搜索的列首
file_path = 'your_file.csv'  # 替换为你的文件路径 (支持 .csv, .xlsx, .xls)

result = __search_by_column(column_name, file_path)

if result:
    # 将结果存储到宿主的列表中
    my_list = result
    print(f"列 '{column_name}' 的所有值已存储到列表: {my_list}")
else:
    print("未找到该列。")